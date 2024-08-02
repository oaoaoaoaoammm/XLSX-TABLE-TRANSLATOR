import pandas as pd
from googletrans import Translator
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import MergedCell

# table name
file_path = 'table_cn.xlsx'
# page name
sheet_name = 'TDSheet Chinese'
# output file name
output_file_path = 'table_cn2.xlsx'

df = pd.read_excel(file_path, sheet_name=sheet_name)

translator = Translator()
#translator.raise_Exception = True


# translate function
def translate_text(text, src='ru', dest='zh-cn'):
    if not text or pd.isna(text):
        return text
    try:
        translated = translator.translate(text, src=src, dest=dest)
        print(f"TRANSLATING: \n {translated.text} \n")
        return translated.text
    except Exception as e:
        print(f"Error: {e}")
        return text

# uploading translated text
for col in df.columns:
    df[col] = df[col].astype(str).apply(lambda x: translate_text(x) if x.strip() else '')

# saving translation
wb = load_workbook(file_path)
ws = wb[sheet_name]

# translate column headers
for i, col in enumerate(ws.iter_cols(1, ws.max_column), 1):
    header_cell = ws.cell(row=1, column=i)
    if not isinstance(header_cell, MergedCell):
        header_cell.value = translate_text(header_cell.value)

# apply translation to other cells
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx)
        if not isinstance(cell, MergedCell):
            cell.value = value

# saving file
wb.save(output_file_path)

print(f"Translation finished. File saved as {output_file_path}")